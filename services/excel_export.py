import asyncio
import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database.models import Product

async def generate_excel_report(products: List[Product], period: str, language: str = "ru") -> str:
    """
    Генерация Excel отчета
    
    Args:
        products: Список товаров
        period: Период отчета ('week', 'month', 'year')
        language: Язык отчета
        
    Returns:
        Путь к файлу
    """
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    
    # Названия колонок в зависимости от языка
    headers = {
        "ru": [
            "№",
            "Трек-код", 
            "Статус",
            "Страна отправления",
            "Тип доставки",
            "Хрупкий",
            "Дата отправки",
            "Ожидаемая доставка",
            "Статус доставки до дверей",
            "Дата создания"
        ],
        "tj": [
            "№",
            "Рамзи тамошобин", 
            "Статус",
            "Кишвари фиристод",
            "Навъи расонидан",
            "Нозук",
            "Санаи фиристод",
            "Расониидани интизоршаванда",
            "Статуси расонидан то дар",
            "Санаи сохт"
        ]
    }
    
    # Заголовки
    ws.append(headers[language])
    
    # Настройка стилей для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Применяем стили к заголовкам
    for col in range(1, len(headers[language]) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
    
    # Заполняем данные
    for i, product in enumerate(products, 2):
        # Перевод статусов
        status_translation = {
            "ru": {
                "created": "Создан",
                "china_warehouse": "На складе Китай",
                "in_transit": "В пути",
                "tajikistan_warehouse": "На складе Таджикистан",
                "delivered": "Доставлен",
                "completed": "Завершен"
            },
            "tj": {
                "created": "Сохта шудааст",
                "china_warehouse": "Дар анбори Чин",
                "in_transit": "Дар роҳ",
                "tajikistan_warehouse": "Дар анбори Тоҷикистон",
                "delivered": "Расонида шуд",
                "completed": "Анҷом ёфт"
            }
        }
        
        # Перевод статусов доставки до дверей
        door_delivery_translation = {
            "ru": {
                "pending": "В ожидании",
                "delivered": "Доставлено",
                "cancelled": "Отменено"
            },
            "tj": {
                "pending": "Дар интизорӣ",
                "delivered": "Расонида шуд",
                "cancelled": "Бекор карда шуд"
            }
        }
        
        row = [
            i - 1,  # №
            product.track_code,
            status_translation[language].get(product.status.value, product.status.value),
            product.country_from or "-",
            product.delivery_type or "-",
            "Да" if product.fragile else "Нет" if language == "ru" else "Ҳа" if product.fragile else "Не",
            product.send_date.strftime("%d.%m.%Y %H:%M") if product.send_date else "-",
            product.expected_delivery_date.strftime("%d.%m.%Y") if product.expected_delivery_date else "-",
            door_delivery_translation[language].get(
                product.door_delivery_status.value, 
                product.door_delivery_status.value
            ),
            product.created_at.strftime("%d.%m.%Y %H:%M")
        ]
        
        ws.append(row)
        
        # Применяем стили к данным
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border
            if col in [1, 7, 8, 10]:  # Центрируем числовые и даты
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Настройка ширины колонок
    column_widths = [5, 20, 15, 15, 15, 10, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Добавляем итоговую строку
    total_row = len(products) + 2
    ws.cell(row=total_row, column=1, value="Итого:" if language == "ru" else "Ҷамъ:")
    ws.cell(row=total_row, column=2, value=len(products))
    
    # Объединяем ячейки для итоговой строки
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=len(headers[language]))
    
    # Стиль для итоговой строки
    total_cell = ws.cell(row=total_row, column=1)
    total_cell.font = Font(bold=True, color="FFFFFF")
    total_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Создаем папку для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)
    
    # Генерируем имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/report_{period}_{timestamp}.xlsx"
    
    # Сохраняем файл
    wb.save(filename)
    
    return filename

async def generate_delivery_report(deliveries: List[Product], language: str = "ru") -> str:
    """
    Генерация отчета по доставке до дверей
    
    Args:
        deliveries: Список товаров с доставкой до дверей
        language: Язык отчета
        
    Returns:
        Путь к файлу
    """
    wb = Workbook()
    ws = wb.active
    
    headers = {
        "ru": [
            "№",
            "Трек-код",
            "Клиент",
            "Телефон",
            "Адрес",
            "Статус доставки",
            "Дата заказа"
        ],
        "tj": [
            "№",
            "Рамзи тамошобин",
            "Мизоҷ",
            "Телефон",
            "Адрес",
            "Статуси расонидан",
            "Сании фармоиш"
        ]
    }
    
    ws.append(headers[language])
    
    # Стили заголовков
    for col in range(1, len(headers[language]) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for i, delivery in enumerate(deliveries, 2):
        row = [
            i - 1,
            delivery.track_code,
            delivery.user.full_name if delivery.user else "-",
            delivery.user.phone if delivery.user else "-",
            delivery.delivery_address or "-",
            delivery.door_delivery_status.value,
            delivery.created_at.strftime("%d.%m.%Y %H:%M")
        ]
        ws.append(row)
    
    # Настройка ширины колонок
    column_widths = [5, 20, 20, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Сохраняем файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/delivery_report_{timestamp}.xlsx"
    wb.save(filename)
    
    return filename